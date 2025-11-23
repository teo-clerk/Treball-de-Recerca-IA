from openpyxl import Workbook
from openpyxl import load_workbook
from os import listdir
import random

ruta = r'E:\TRABAJO DE RECERCA\Treball-de-Recerca-IA\TR\DATOS'
arxius = listdir(ruta)
generaciones = 125
print(listdir(ruta))
for nombre in arxius:
    with open(r"E:\TRABAJO DE RECERCA\Treball-de-Recerca-IA\TR\DATOS\\" + nombre) as f:
        contents = f.readlines()
        matrix = []
        for i in range(len(contents)):
            matrix.append([eval(i) for i in (contents[i].replace("\n", "").replace(",", "").split())])
        matrix.pop(-1)
    #acceso al archivo .txt i  lectura de datos para una posterior transformación  en una matriz

    #completar datos/matriz
    #print("partidas:   " + str(len(matrix)))
    for i in range(len(matrix)):
        maxim = max(matrix[i])
        for j in range(generaciones - len(matrix[i])):
            matrix[i].append(maxim)

        #print(len(matrix[i]))

    #print(matrix)
            
    lista_promedio = []
    for i in range(len(matrix[0])):
        n = 0
        for f in range(len(matrix)):
            n += matrix[f][i]
        lista_promedio.append(round((n/len(matrix)), 2))
    #creación de una lista que contiene un promedio de todos los valores

    wb = load_workbook(filename = 'IA2.xlsx')
    #print(wb.sheetnames)

    columnas_excel = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"]
    #workbook = Workbook()
    #sheet = workbook.active
    
    #definicion de variables

    nw = wb["Sheet"]
    if nw["A1"].value != None:
        #name = "new" + str(random.randint(1,1000))

        name = nombre.replace('[', '').replace(']', '').replace('oblació', '').replace('nputs', '').replace('utacions', '').replace('PY_', '').replace('_0.txt', '')
        new = wb.create_sheet(name)
        new = wb[name]
        for i in range(len(lista_promedio)):
            new[columnas_excel[0] + str(i+1)] = i+1
            new[columnas_excel[1] + str(i+1)] = lista_promedio[i]
    #creación de una tabla en un excel con los valores medios de fitness de cada generacion
    else:
        for i in range(len(lista_promedio)):
            nw[columnas_excel[0] + str(i+1)] = i+1
            nw[columnas_excel[1] + str(i+1)] = lista_promedio[i]
        
    #creación de una nueva hoja en caso de  que la hoja activa no esté vacía

    wb.save(filename="IA2.xlsx")
    #guardado del archivo

    """
        from openpyxl import load_workbook
        wb = load_workbook(filename = 'empty_book.xlsx')
        sheet_ranges = wb['range names']
        print(sheet_ranges['D18'].value)
        3
    """